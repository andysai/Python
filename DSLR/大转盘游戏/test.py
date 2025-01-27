import requests
import json
import openpyxl

# 高德地图行政区划API的URL，需要替换为实际的API Key和层级代码（例如全国为100000）
api_key = '84447d8cea0cb7c6d56156998019041c'
base_url = f"https://restapi.amap.com/v3/config/district?subdistrict=3&extensions=base&key={api_key}"

response = requests.get(base_url)
data = response.json()
wb = openpyxl.Workbook()
ws = wb.active

if data['status'] == '1':  # 检查状态码是否表示成功
    provinces = data['districts'][0]['districts']  # 获取省份列表，层级为3（省、直辖市、自治区）
    # print(provinces)
    for province in provinces:
        # print('区号编码：' + province['adcode'] + '\n' + '省份名称：' + province['name'])

        for city in province['districts']:
            province_name = province['name']
            city_name = city['name']
            full_name = province_name + '-' + city_name
            print(len(full_name))
            # for long_list in len(full_name):
            #     ws['A' + long_list] = province_name
            #     ws['B1'] = city_name
            #     ws['C1'] = full_name
else:
    print("获取数据失败")

wb.save('test.xlsx')
wb.close()
